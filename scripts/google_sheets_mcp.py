import io
import os
import sys
from pathlib import Path

# Thêm dependencies vào virtualenv nếu chưa có
try:
    from mcp.server.fastmcp import FastMCP
    import openpyxl
    from google.oauth2.service_account import Credentials
    from google.auth.transport.requests import AuthorizedSession
except ImportError:
    print("❌ Lỗi: Thiếu thư viện. Vui lòng chạy lệnh sau để cài đặt:")
    print("   pip install mcp openpyxl google-auth requests")
    sys.exit(1)

# Khởi tạo FastMCP Server
mcp = FastMCP("SRRM Google Sheets Task Manager")

FILE_ID = "1z-SdPWU9hosISB2T71sYZ03-x44OzGvv"
CREDENTIALS_FILE = Path(__file__).parent.parent / "service_account.json"

@mcp.tool()
def update_task_status(task_id: str, status: str) -> str:
    """
    Cập nhật trạng thái của một Task trong Google Sheet (định dạng Excel .xlsx tải lên Drive) của dự án.
    
    Args:
        task_id: Mã định danh của Task (Ví dụ: 'AI-01.1', 'SRRM-D1').
        status: Trạng thái mới (Ví dụ: 'Completed'/'Hoàn thành', 'In Progress'/'Đang làm', 'Pending'/'Chưa làm', 'Blocked'/'Bị chặn').
    """
    if not CREDENTIALS_FILE.exists():
        return (
            f"Error: Chưa tìm thấy tệp thông tin xác thực tại: {CREDENTIALS_FILE.absolute()}\n"
            "Vui lòng tạo Service Account trên Google Cloud Console, tải tệp JSON key và đổi tên thành 'service_account.json' ở gốc dự án."
        )

    try:
        # Sử dụng Drive API scope để có thể tải và cập nhật file Excel gốc
        scopes = ["https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(str(CREDENTIALS_FILE), scopes=scopes)
        authed_session = AuthorizedSession(creds)
        
        # 1. Download file XLSX
        download_url = f"https://www.googleapis.com/drive/v3/files/{FILE_ID}?alt=media"
        response = authed_session.get(download_url)
        if response.status_code != 200:
            return f"Error: Không thể tải file từ Google Drive. Status code: {response.status_code}, Detail: {response.text}"
            
        # 2. Đọc file bằng openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        
        # Tự động tìm hàng tiêu đề (Header row)
        header_row = None
        task_id_col = None
        status_col = None
        
        for r in range(1, 20):  # Quét 20 dòng đầu
            row_vals = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[r]]
            if any(val in ["mã", "task id", "task_id"] for val in row_vals) and any(val in ["trạng thái", "status", "state"] for val in row_vals):
                header_row = r
                for c_idx, val in enumerate(row_vals, 1):
                    if val in ["mã", "task id", "task_id"]:
                        task_id_col = c_idx
                    elif val in ["trạng thái", "status", "state"]:
                        status_col = c_idx
                break
                
        if not header_row or not task_id_col or not status_col:
            return "Error: Không tự động tìm thấy dòng tiêu đề chứa 'Mã' và 'Trạng thái' trong trang tính."
            
        # Tìm hàng chứa Task ID tương ứng
        target_row = None
        for r in range(header_row + 1, sheet.max_row + 1):
            val = sheet.cell(row=r, column=task_id_col).value
            if val and str(val).strip() == task_id.strip():
                target_row = r
                break
                
        if not target_row:
            return f"Error: Không tìm thấy Task ID '{task_id}' trong danh sách."
            
        # Ánh xạ trạng thái sang Tiếng Việt tương thích với dropdown của Sheet
        status_mapping = {
            "completed": "Hoàn thành",
            "done": "Hoàn thành",
            "hoàn thành": "Hoàn thành",
            "in progress": "Đang làm",
            "doing": "Đang làm",
            "đang làm": "Đang làm",
            "pending": "Chưa làm",
            "todo": "Chưa làm",
            "chưa làm": "Chưa làm",
            "blocked": "Bị chặn",
            "bị chặn": "Bị chặn"
        }
        
        mapped_status = status_mapping.get(status.strip().lower(), status)
        
        # Cập nhật ô
        sheet.cell(row=target_row, column=status_col, value=mapped_status)
        
        # 3. Ghi đè file lên Google Drive
        out_bytes = io.BytesIO()
        wb.save(out_bytes)
        out_bytes.seek(0)
        xlsx_data = out_bytes.read()
        
        upload_url = f"https://www.googleapis.com/upload/drive/v3/files/{FILE_ID}?uploadType=media"
        headers = {"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        
        upload_resp = authed_session.patch(upload_url, data=xlsx_data, headers=headers)
        if upload_resp.status_code == 200:
            return f"✅ Đã cập nhật trạng thái của Task {task_id} thành '{mapped_status}' trên Google Drive thành công!"
        else:
            return f"Error: Không thể lưu file ghi đè lên Google Drive. Status code: {upload_resp.status_code}, Detail: {upload_resp.text}"
            
    except Exception as e:
        return f"Error: Đã xảy ra lỗi khi kết nối Google Drive API: {str(e)}"

if __name__ == "__main__":
    mcp.run()
